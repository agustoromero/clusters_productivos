#!/usr/bin/env python3
"""Ejecución end-to-end para salarios con validaciones de calidad (sin dependencias externas)."""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
from collections import defaultdict
from pathlib import Path
from typing import Dict, Iterable, List, Tuple


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()
    p.add_argument("--salarios", required=True, help="Ruta CSV salarios mensual")
    p.add_argument("--ipc", required=True, help="Ruta IPC (csv/xlsx)")
    p.add_argument("--outdir", default="outputs", help="Carpeta de salida")
    p.add_argument("--anio", type=int, default=None, help="Filtrar procesamiento a un único año (ej: 2022)")
    p.add_argument(
        "--sectores-esperados",
        default="A,B,C,D,E,F,G",
        help="Lista de letras esperadas separadas por coma para controlar cobertura sectorial",
    )
    return p.parse_args()


def to_float(v: str) -> float:
    if v is None:
        return math.nan
    s = str(v).strip().replace(",", ".")
    if s == "":
        return math.nan
    try:
        return float(s)
    except ValueError:
        return math.nan


def zfill_periodo(v: str) -> str:
    digits = re.sub(r"\D", "", str(v))
    return digits.zfill(6)[:6]


def read_csv_rows(path: Path) -> List[Dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return [{(k or "").strip().lower(): (v or "") for k, v in row.items()} for row in reader]


def read_ipc(path: Path) -> Dict[str, Dict[str, float]]:
    if path.suffix.lower() == ".csv":
        rows = read_csv_rows(path)
    else:
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception as e:
            raise RuntimeError("Para IPC en xlsx necesitás openpyxl o exportar IPC a CSV") from e
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: "" if r[i] is None else str(r[i]) for i in range(len(headers))})

    out: Dict[str, Dict[str, float]] = {}
    for row in rows:
        per = zfill_periodo(row.get("periodo", ""))
        if not per.strip("0"):
            continue
        out[per] = {
            "IPC": to_float(row.get("ipc", row.get("IPC", ""))),
            "INF": to_float(row.get("inf", row.get("INF", ""))),
        }
    return out


def normalize_sector_to_2d(text: str) -> str:
    s = (text or "").strip().lower()
    if not s:
        return ""
    m = re.search(r"\b([a-z])\b", s)
    if m:
        return m.group(1).upper()
    if "serv" in s:
        return "SERV"
    return s[:2].upper()


def threshold_status(p: float) -> str:
    if p < 0.05:
        return "ok"
    if p <= 0.20:
        return "warning"
    return "error"


def safe_mean(values: Iterable[float]) -> float:
    vals = [v for v in values if not math.isnan(v)]
    if not vals:
        return math.nan
    return sum(vals) / len(vals)


def write_csv(path: Path, rows: List[Dict[str, object]], columns: List[str]) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=columns)
        w.writeheader()
        for r in rows:
            w.writerow(r)


def main() -> None:
    args = parse_args()
    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    sal_rows = read_csv_rows(Path(args.salarios))
    ipc = read_ipc(Path(args.ipc))
    sectores_esperados = [s.strip().upper() for s in str(args.sectores_esperados).split(",") if s.strip()]

    required = {"periodo", "id_prov", "sector", "salarios"}
    if not sal_rows:
        raise RuntimeError("CSV salarios sin filas")
    missing_cols = [c for c in required if c not in sal_rows[0]]
    if missing_cols:
        raise RuntimeError(f"Faltan columnas en salarios: {missing_cols}")

    monthly: List[Dict[str, object]] = []
    for r in sal_rows:
        periodo = zfill_periodo(r.get("periodo", ""))
        anio = int(periodo[:4]) if periodo[:4].isdigit() else None
        mes = int(periodo[4:6]) if periodo[4:6].isdigit() else None
        if mes is None or mes < 1 or mes > 12:
            continue
        if args.anio is not None and anio != args.anio:
            continue

        salarios = to_float(r.get("salarios", ""))
        ip = ipc.get(periodo, {})
        ipc_val = ip.get("IPC", math.nan)
        salario_const = salarios * (100.0 / ipc_val) if (not math.isnan(salarios) and not math.isnan(ipc_val) and ipc_val > 0) else math.nan
        empleo = to_float(r.get("empleo", ""))

        monthly.append(
            {
                "periodo": periodo,
                "anio": anio,
                "mes": mes,
                "id_prov": str(r.get("id_prov", "")).strip(),
                "sector": r.get("sector", ""),
                "actividad_2d": normalize_sector_to_2d(r.get("sector", "")),
                "salarios": salarios,
                "IPC": ipc_val,
                "INF": ip.get("INF", math.nan),
                "salario_const": salario_const,
                "empleo": empleo,
                "salario_missing": math.isnan(salario_const),
                "ipc_missing": math.isnan(ipc_val) or ipc_val <= 0,
                "merge_salario_fallido": math.isnan(salarios),
            }
        )

    empleo_sum: Dict[Tuple[str, str, int], float] = defaultdict(float)
    empleo_count: Dict[Tuple[str, str, int], int] = defaultdict(int)
    for r in monthly:
        key = (r["id_prov"], r["actividad_2d"], r["anio"])
        if not math.isnan(r["empleo"]):
            empleo_sum[key] += r["empleo"]
            empleo_count[key] += 1

    for r in monthly:
        key = (r["id_prov"], r["actividad_2d"], r["anio"])
        esum = empleo_sum.get(key, 0.0)
        r["empleo_sum"] = esum
        r["empleo_sum_cero"] = empleo_count.get(key, 0) > 0 and esum <= 0

    grouped: Dict[Tuple[str, str, int], List[Dict[str, object]]] = defaultdict(list)
    for r in monthly:
        grouped[(r["id_prov"], r["actividad_2d"], r["anio"])] .append(r)

    annual: List[Dict[str, object]] = []
    for (id_prov, actividad_2d, anio), rows in grouped.items():
        sal_const_vals = [x["salario_const"] for x in rows if not math.isnan(x["salario_const"])]
        n_meses = len(sal_const_vals)
        promedio = safe_mean(sal_const_vals)
        numerador_pond = 0.0
        denominador_pond = 0.0
        for x in rows:
            sc = x["salario_const"]
            emp = x["empleo"]
            if not math.isnan(sc) and not math.isnan(emp) and emp > 0:
                numerador_pond += sc * emp
                denominador_pond += emp
        promedio_pond = (numerador_pond / denominador_pond) if denominador_pond > 0 else math.nan
        row = {
            "id_prov": id_prov,
            "actividad_2d": actividad_2d,
            "anio": anio,
            "n_meses_observados": n_meses,
            "salario_promedio_anual_const": promedio,
            "salario_promedio_anual_pond_const": promedio_pond,
            "denominador_pond_anual": denominador_pond,
            "flag_cobertura_mensual_baja": n_meses < 6,
            "flag_ponderado_sin_muestra": denominador_pond <= 0,
        }
        if n_meses >= 6:
            annual.append(row)

    # Diagnóstico sector-año para detectar ceros y baja muestra.
    sector_anio: Dict[Tuple[int, str], Dict[str, object]] = {}
    for r in monthly:
        key = (r["anio"], r["actividad_2d"])
        if key not in sector_anio:
            sector_anio[key] = {
                "anio": r["anio"],
                "actividad_2d": r["actividad_2d"],
                "n_filas": 0,
                "n_salario_const_validos": 0,
                "n_salario_const_cero": 0,
                "periodos": set(),
                "provincias": set(),
            }
        info = sector_anio[key]
        info["n_filas"] += 1
        info["periodos"].add(r["periodo"])
        info["provincias"].add(r["id_prov"])
        if not math.isnan(r["salario_const"]):
            info["n_salario_const_validos"] += 1
            if r["salario_const"] == 0:
                info["n_salario_const_cero"] += 1

    sector_anio_rows: List[Dict[str, object]] = []
    for _, info in sorted(sector_anio.items(), key=lambda x: (x[0][0], x[0][1])):
        n_valid = info["n_salario_const_validos"]
        n_zero = info["n_salario_const_cero"]
        pct_zero = (n_zero / n_valid) if n_valid else math.nan
        row = {
            "anio": info["anio"],
            "actividad_2d": info["actividad_2d"],
            "n_filas": info["n_filas"],
            "n_salario_const_validos": n_valid,
            "n_salario_const_cero": n_zero,
            "pct_salario_const_cero": pct_zero,
            "n_periodos_distintos": len(info["periodos"]),
            "n_provincias_distintas": len(info["provincias"]),
            "flag_muestra_baja": len(info["periodos"]) < 6,
        }
        sector_anio_rows.append(row)

    total = len(monthly) or 1
    pct_ceros = sum(1 for r in monthly if r["salario_const"] == 0) / total
    pct_nan_pre = sum(1 for r in monthly if math.isnan(r["salario_const"])) / total
    pct_nan_post_merge = sum(1 for r in monthly if math.isnan(r["salarios"])) / total
    rows_with_empleo = [r for r in monthly if not math.isnan(r["empleo"])]
    pct_empleo_sum_cero = (
        sum(1 for r in rows_with_empleo if r["empleo_sum_cero"]) / (len(rows_with_empleo) or 1)
    )

    quality = {
        "salarios": {
            "pct_ceros_salario_real": pct_ceros,
            "pct_ceros_status": threshold_status(pct_ceros),
            "pct_nan_salario_real_pre_fill": pct_nan_pre,
            "pct_nan_salario_nominal_post_merge": pct_nan_post_merge,
            "pct_grupos_empleo_sum_cero": pct_empleo_sum_cero,
            "sectores_cubiertos": len({r["actividad_2d"] for r in monthly if r["actividad_2d"]}),
            "anios_cubiertos": sorted({r["anio"] for r in monthly if r["anio"] is not None}),
            "ipc_min": None if not monthly else min((r["IPC"] for r in monthly if not math.isnan(r["IPC"])), default=None),
            "ipc_nulos": sum(1 for r in monthly if math.isnan(r["IPC"])),
            "filas_mensuales": len(monthly),
            "filas_anuales_validas": len(annual),
            "sectores_esperados": sectores_esperados,
        }
    }
    if args.anio is not None:
        sectores_presentes = sorted(
            {r["actividad_2d"] for r in sector_anio_rows if r["anio"] == args.anio and r["actividad_2d"]}
        )
        sectores_faltantes = sorted([s for s in sectores_esperados if s not in sectores_presentes])
        sectores_anio_cero = sorted(
            {
                r["actividad_2d"]
                for r in annual
                if r["anio"] == args.anio and not math.isnan(r["salario_promedio_anual_const"]) and r["salario_promedio_anual_const"] == 0
            }
        )
        sectores_baja_muestra = sorted(
            {
                r["actividad_2d"]
                for r in sector_anio_rows
                if r["anio"] == args.anio and r["flag_muestra_baja"]
            }
        )
        quality["salarios"]["anio_objetivo"] = args.anio
        quality["salarios"]["sectores_presentes_anio_objetivo"] = sectores_presentes
        quality["salarios"]["sectores_faltantes_anio_objetivo"] = sectores_faltantes
        quality["salarios"]["sectores_con_salario_anual_cero_anio_objetivo"] = sectores_anio_cero
        quality["salarios"]["sectores_con_baja_muestra_anio_objetivo"] = sectores_baja_muestra

    write_csv(
        outdir / "salarios_mensual_clean.csv",
        monthly,
        [
            "periodo",
            "anio",
            "mes",
            "id_prov",
            "sector",
            "actividad_2d",
            "salarios",
            "IPC",
            "INF",
            "salario_const",
            "empleo",
            "empleo_sum",
            "salario_missing",
            "ipc_missing",
            "merge_salario_fallido",
            "empleo_sum_cero",
        ],
    )
    write_csv(
        outdir / "salarios_anual_prov_actividad2d.csv",
        annual,
        [
            "id_prov",
            "actividad_2d",
            "anio",
            "n_meses_observados",
            "salario_promedio_anual_const",
            "salario_promedio_anual_pond_const",
            "denominador_pond_anual",
            "flag_cobertura_mensual_baja",
            "flag_ponderado_sin_muestra",
        ],
    )
    write_csv(
        outdir / "quality_sector_anio.csv",
        sector_anio_rows,
        [
            "anio",
            "actividad_2d",
            "n_filas",
            "n_salario_const_validos",
            "n_salario_const_cero",
            "pct_salario_const_cero",
            "n_periodos_distintos",
            "n_provincias_distintas",
            "flag_muestra_baja",
        ],
    )
    (outdir / "quality_ingesta.json").write_text(json.dumps(quality, ensure_ascii=False, indent=2), encoding="utf-8")

    print("=== RESUMEN E2E ===")
    print(json.dumps(quality, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
